"""Form 1030 as a live .xlsx — one sheet per year, plus Cumulative and Flags.

Subtotals are FORMULAS, not baked values: ORA changes one cell and the sheet
re-totals. That is the whole reason Excel was chosen over a static PDF.

Design: docs/superpowers/specs/2026-09-03-nsf-form-1030-budget-design.md
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from services.nsf_budget import G_ITEM_LINES, OTHER_PERSONNEL_ROWS, compute_document

MONEY = '"$"#,##0'
_HEAD_FILL = PatternFill("solid", fgColor="0B2F5E")
_TOTAL_FILL = PatternFill("solid", fgColor="EEF2F7")
_HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
_BOLD = Font(bold=True)


def _money_cell(ws, row, value):
    c = ws.cell(row=row, column=4, value=value)
    c.number_format = MONEY
    return c


def _label(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=1, value=text)
    c.alignment = Alignment(indent=indent)
    if bold:
        c.font = _BOLD
    return c


def _write_year(ws, meta, yc):
    """Lay one A-M year sheet out. Column A is the label, column D the money."""
    L = yc["lines"]
    ws.column_dimensions["A"].width = 56
    for col in "BCD":
        ws.column_dimensions[col].width = 16

    ws["A1"] = "SUMMARY PROPOSAL BUDGET"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Organization: {meta.get('organization', '')}"
    ws["A3"] = f"Principal Investigator / Project Director: {meta.get('pi_name', '')}"
    ws["A4"] = f"Duration (months): {meta.get('duration_months', '')}"
    ws["A5"] = (f"Year {yc['year']}" if isinstance(yc["year"], int)
                else str(yc["year"]))
    ws["A5"].font = _BOLD

    r = 7
    for col in range(1, 5):
        ws.cell(row=r, column=col).fill = _HEAD_FILL
    c = ws.cell(row=r, column=1, value="NSF-FUNDED PERSON-MONTHS / FUNDS REQUESTED")
    c.font = _HEAD_FONT
    ws.cell(row=r, column=2, value="MONTHS").font = _HEAD_FONT
    ws.cell(row=r, column=4, value="FUNDS REQUESTED").font = _HEAD_FONT

    r += 1
    _label(ws, r, "A. SENIOR/KEY PERSONNEL", bold=True)
    a_first = r + 1
    for row in L["A"]["rows"]:
        r += 1
        _label(ws, r, f"{row['name']}" + (f" - {row['role']}" if row["role"] else ""),
               indent=1)
        ws.cell(row=r, column=2, value=row["months_total"])
        _money_cell(ws, r, row["salary"])
    if r < a_first:                       # cumulative sheet carries no rows
        r += 1
        _label(ws, r, "(see the year sheets)", indent=1)
        _money_cell(ws, r, L["A"]["total"])
    a_last = r
    r += 1
    _label(ws, r, "TOTAL SENIOR/KEY PERSONNEL", indent=1, bold=True)
    _money_cell(ws, r, f"=SUM(D{a_first}:D{a_last})").font = _BOLD
    a_total_row = r

    r += 1
    _label(ws, r, "B. OTHER PERSONNEL", bold=True)
    b_first = r + 1
    b_by_key = {x["key"]: x for x in L["B"]["rows"]} if L["B"]["rows"] else {}
    for key, label, _has_months in OTHER_PERSONNEL_ROWS:
        r += 1
        row = b_by_key.get(key)
        count = row["count"] if row else 0
        _label(ws, r, f"({count}) {label}", indent=1)
        _money_cell(ws, r, row["amount"] if row else 0.0)
    b_last = r
    r += 1
    _label(ws, r, "TOTAL SALARIES AND WAGES (A + B)", indent=1, bold=True)
    _money_cell(ws, r, f"=D{a_total_row}+SUM(D{b_first}:D{b_last})").font = _BOLD
    ab_row = r

    r += 2
    _label(ws, r, "C. FRINGE BENEFITS (IF CHARGED AS DIRECT COSTS)", bold=True)
    _money_cell(ws, r, L["C"])
    c_row = r
    r += 1
    _label(ws, r, "TOTAL SALARIES, WAGES AND FRINGE BENEFITS (A + B + C)",
           indent=1, bold=True)
    _money_cell(ws, r, f"=D{ab_row}+D{c_row}").font = _BOLD
    abc_row = r

    r += 2
    _label(ws, r, "D. EQUIPMENT (LIST EACH ITEM EXCEEDING $10,000)", bold=True)
    d_first = r + 1
    for item in L["D"]["rows"]:
        r += 1
        _label(ws, r, item["description"] or "Unnamed item", indent=1)
        _money_cell(ws, r, item["amount"])
    if r < d_first:
        r += 1
        _label(ws, r, "(see the year sheets)", indent=1)
        _money_cell(ws, r, L["D"]["total"])
    d_last = r
    r += 1
    _label(ws, r, "TOTAL EQUIPMENT", indent=1, bold=True)
    _money_cell(ws, r, f"=SUM(D{d_first}:D{d_last})").font = _BOLD
    d_row = r

    r += 2
    _label(ws, r, "E. TRAVEL", bold=True)
    r += 1
    _label(ws, r, "1. DOMESTIC (INCL. U.S. POSSESSIONS)", indent=1)
    _money_cell(ws, r, L["E"]["domestic"])
    e_first = r
    r += 1
    _label(ws, r, "2. INTERNATIONAL", indent=1)
    _money_cell(ws, r, L["E"]["international"])
    e_last = r
    r += 1
    _label(ws, r, "TOTAL TRAVEL", indent=1, bold=True)
    _money_cell(ws, r, f"=SUM(D{e_first}:D{e_last})").font = _BOLD
    e_row = r

    r += 2
    _label(ws, r, "F. PARTICIPANT SUPPORT COSTS", bold=True)
    f_first = r + 1
    for key, label in (("stipends", "1. STIPENDS"), ("travel", "2. TRAVEL"),
                       ("subsistence", "3. SUBSISTENCE"), ("other", "4. OTHER")):
        r += 1
        _label(ws, r, label, indent=1)
        _money_cell(ws, r, L["F"][key])
    f_last = r
    r += 1
    _label(ws, r, f"TOTAL NUMBER OF PARTICIPANTS ({L['F']['count']})"
                  f"   TOTAL PARTICIPANT COSTS", indent=1, bold=True)
    _money_cell(ws, r, f"=SUM(D{f_first}:D{f_last})").font = _BOLD
    f_row = r

    r += 2
    _label(ws, r, "G. OTHER DIRECT COSTS", bold=True)
    g_first = r + 1
    for n, (key, label) in enumerate(G_ITEM_LINES, start=1):
        r += 1
        _label(ws, r, f"{n}. {label.upper()}", indent=1)
        _money_cell(ws, r, L["G"][key])
    r += 1
    _label(ws, r, "5. SUBAWARDS", indent=1)
    _money_cell(ws, r, L["G"]["subawards"]["total"])
    r += 1
    _label(ws, r, "6. OTHER", indent=1)
    _money_cell(ws, r, L["G"]["other"])
    g_last = r
    r += 1
    _label(ws, r, "TOTAL OTHER DIRECT COSTS", indent=1, bold=True)
    _money_cell(ws, r, f"=SUM(D{g_first}:D{g_last})").font = _BOLD
    g_row = r

    r += 2
    _label(ws, r, "H. TOTAL DIRECT COSTS (A THROUGH G)", bold=True)
    ws.cell(row=r, column=1).fill = _TOTAL_FILL
    _money_cell(ws, r, f"=D{abc_row}+D{d_row}+D{e_row}+D{f_row}+D{g_row}").font = _BOLD
    h_row = r

    r += 1
    _label(ws, r, f"I. INDIRECT COSTS (F&A) - {yc['fa']['label']} at "
                  f"{yc['fa']['rate'] * 100:.0f}% of an MTDC base of "
                  f"{yc['mtdc']['base']:,.0f}", bold=True)
    _money_cell(ws, r, L["I"]).font = _BOLD
    i_row = r

    r += 1
    _label(ws, r, "J. TOTAL DIRECT AND INDIRECT COSTS (H + I)", bold=True)
    _money_cell(ws, r, f"=D{h_row}+D{i_row}").font = _BOLD
    j_row = r

    r += 1
    _label(ws, r, "K. FEE", bold=True)
    _money_cell(ws, r, L["K"])
    k_row = r

    r += 1
    _label(ws, r, "L. AMOUNT OF THIS REQUEST (J MINUS K)", bold=True)
    ws.cell(row=r, column=1).fill = _TOTAL_FILL
    _money_cell(ws, r, f"=D{j_row}-D{k_row}").font = _BOLD

    r += 1
    _label(ws, r, "M. COST SHARING PROPOSED LEVEL", bold=True)
    _money_cell(ws, r, L["M"])

    r += 2
    ws.cell(row=r, column=1,
            value="MTDC excludes equipment, participant support, the portion of each "
                  "subaward over $25,000, and items marked exempt (tuition remission, "
                  "scholarships, rent, patient care).").font = Font(italic=True, size=9)


def build_workbook(doc, computed=None):
    """A Workbook with one sheet per year, plus Cumulative and Flags."""
    computed = computed or compute_document(doc)
    meta = computed.get("meta") or {}

    wb = Workbook()
    wb.remove(wb.active)

    for yc in computed["years"]:
        _write_year(wb.create_sheet(f"Year {yc['year']}"), meta, yc)

    cum = dict(computed["cumulative"])
    cum["year"] = "Cumulative (all years)"
    cum.setdefault("fa", computed["years"][0]["fa"] if computed["years"]
                   else {"label": "", "rate": 0.0})
    _write_year(wb.create_sheet("Cumulative"), meta, cum)

    fws = wb.create_sheet("Flags")
    for col, (head, width) in enumerate(
            [("YEAR", 10), ("LINE", 8), ("FINDING", 44),
             ("DETAIL", 62), ("CITATION", 26)], start=1):
        c = fws.cell(row=1, column=col, value=head)
        c.fill, c.font = _HEAD_FILL, _HEAD_FONT
        fws.column_dimensions[c.column_letter].width = width
    for n, f in enumerate(computed.get("flags") or [], start=2):
        fws.cell(row=n, column=1, value=f.get("year") or "all")
        fws.cell(row=n, column=2, value=f["line"])
        fws.cell(row=n, column=3, value=f"[{f['severity'].upper()}] {f['title']}")
        fws.cell(row=n, column=4, value=f.get("detail") or f["message"])
        fws.cell(row=n, column=5, value=f["citation"])
    return wb


def workbook_bytes(doc, computed=None):
    """The workbook as bytes, ready to stream from an endpoint."""
    buf = io.BytesIO()
    build_workbook(doc, computed).save(buf)
    return buf.getvalue()
