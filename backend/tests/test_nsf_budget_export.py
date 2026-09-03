"""The .xlsx export must be a LIVE workbook, not a picture of one -- ORA edits
a cell and the sheet re-totals. So subtotals are formulas, not baked values."""
import io

from openpyxl import load_workbook

from services import nsf_budget as nb
from services import nsf_budget_export as ex


def _doc():
    doc = nb.blank_document(years=2)
    for y in doc["years"]:
        y["senior"][0].update(name="Dr. Oladunni", base_salary=90_000,
                              appointment_basis="academic_9", acad=2)
        y["equipment"] = [{"description": "Confocal", "amount": 40_000}]
    return doc


def _loaded(doc, **kw):
    return load_workbook(io.BytesIO(ex.workbook_bytes(doc)), **kw)


def _text(ws):
    return " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)


def test_workbook_has_one_sheet_per_year_plus_cumulative_and_flags():
    wb = _loaded(_doc())
    assert wb.sheetnames == ["Year 1", "Year 2", "Cumulative", "Flags"]


def test_subtotals_are_formulas_not_baked_values():
    wb = _loaded(_doc())          # data_only=False is the default
    ws = wb["Year 1"]
    formulas = [c.value for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=")]
    assert any(f.startswith("=SUM(") for f in formulas)


def test_the_header_block_carries_the_organization_and_pi():
    doc = _doc()
    doc["meta"]["pi_name"] = "Timothy Oladunni"
    text = _text(_loaded(doc)["Year 1"])
    assert "Morgan State University" in text
    assert "Timothy Oladunni" in text


def test_every_form_line_letter_appears_on_a_year_sheet():
    text = _text(_loaded(_doc())["Year 1"])
    for label in ["A.", "B.", "C.", "D.", "E.", "F.", "G.",
                  "H.", "I.", "J.", "K.", "L.", "M."]:
        assert label in text, f"missing form line {label}"


def test_the_flags_sheet_lists_findings_with_their_citation():
    doc = _doc()
    doc["years"][0]["fee"] = 5_000
    text = _text(_loaded(doc)["Flags"])
    assert "PAPPG 24-1 II.D.2.f(x)" in text


def test_a_blank_budget_still_exports():
    assert len(ex.workbook_bytes(nb.blank_document())) > 0


# --- the workbook's own arithmetic must agree with the backend -------------

def _resolve(ws, cell_ref, _depth=0):
    """Evaluate the small formula grammar this exporter emits.

    Wrong cell references are the main failure mode of a formula-based export,
    and openpyxl reads back the formula, not a value -- so we evaluate it.
    Supports: a literal, =SUM(Da:Db), and +/- chains of terms.
    """
    assert _depth < 25, "formula recursion"
    v = ws[cell_ref].value
    if not isinstance(v, str) or not v.startswith("="):
        return float(v or 0)

    total, sign = 0.0, 1
    term = ""
    for ch in v[1:] + "+":
        if ch in "+-":
            term = term.strip()
            if term.upper().startswith("SUM("):
                start, end = term[4:-1].split(":")
                col = "".join(c for c in start if c.isalpha())
                lo = int("".join(c for c in start if c.isdigit()))
                hi = int("".join(c for c in end if c.isdigit()))
                total += sign * sum(_resolve(ws, f"{col}{n}", _depth + 1)
                                    for n in range(lo, hi + 1))
            elif term:
                total += sign * _resolve(ws, term, _depth + 1)
            sign, term = (1 if ch == "+" else -1), ""
        else:
            term += ch
    return round(total, 2)


def _find_row(ws, prefix):
    for row in ws.iter_rows(min_col=1, max_col=1):
        if isinstance(row[0].value, str) and row[0].value.startswith(prefix):
            return row[0].row
    raise AssertionError(f"row starting {prefix!r} not found")


def test_the_worksheet_formulas_reproduce_the_computed_totals():
    """The worked example: H 196,100 / L 247,994, computed by the SHEET."""
    doc = nb.blank_document()
    s = doc["years"][0]
    s["senior"][0].update(name="Dr. Oladunni", base_salary=90_000,
                          appointment_basis="academic_9", acad=2)
    s["other_personnel"]["grad_students"] = {
        "count": 1, "amount": 30_000, "fringe_key": "contractual"}
    s["equipment"] = [{"description": "Confocal", "amount": 40_000}]
    s["travel"]["domestic"] = [{"description": "Conf", "amount": 3_000}]
    s["travel"]["international"] = [{"description": "Collab", "amount": 2_000}]
    s["participant_support"] = {"count": 15, "stipends": 10_000, "travel": None,
                                "subsistence": None, "other": None}
    s["other_direct"]["materials_supplies"] = [{"description": "Reagents", "amount": 5_000}]
    s["other_direct"]["subawards"] = [{"organization": "Partner U", "amount": 50_000}]
    s["other_direct"]["other"] = [
        {"description": "Grad tuition remission", "amount": 25_000, "mtdc_exempt": True}]

    computed = nb.compute_document(doc)
    ws = _loaded(doc)["Year 1"]

    assert _resolve(ws, f"D{_find_row(ws, 'H.')}") == computed["years"][0]["lines"]["H"]
    assert _resolve(ws, f"D{_find_row(ws, 'J.')}") == computed["years"][0]["lines"]["J"]
    assert _resolve(ws, f"D{_find_row(ws, 'L.')}") == computed["years"][0]["lines"]["L"]
    assert _resolve(ws, f"D{_find_row(ws, 'H.')}") == 196_100.0
    assert _resolve(ws, f"D{_find_row(ws, 'L.')}") == 247_994.0


def test_a_fee_flows_through_the_sheet_formula_for_line_l():
    doc = nb.blank_document()
    doc["years"][0]["other_direct"]["materials_supplies"] = [
        {"description": "Reagents", "amount": 10_000}]
    doc["years"][0]["fee"] = 1_000
    ws = _loaded(doc)["Year 1"]
    assert _resolve(ws, f"D{_find_row(ws, 'J.')}") == 15_400.0
    assert _resolve(ws, f"D{_find_row(ws, 'L.')}") == 14_400.0     # J - K
